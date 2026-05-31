#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
文人·外贸客户号码查找工具 - 通过搜索引擎查找产品关键词对应的企业网站，提取联系信息。
支持：DuckDuckGo / Google API / Bing API · 多语言关键词 · 自动过滤非企业网站

用法：
    python number_finder.py "LED lights" -c de           # 谷歌德国站搜索
    python number_finder.py "acero inoxidable" -c es -e bing  # 必应西班牙站
    python number_finder.py "moteur électrique" -c fr -n 30    # 最多30条结果
    python number_finder.py "LED lights" -c de uk fr           # 多国搜索
"""

import re
import sys
import csv
import time
import random
import hashlib
import logging
import urllib.parse
from pathlib import Path
from datetime import datetime
from collections import OrderedDict

import requests
from bs4 import BeautifulSoup

# ── Windows终端UTF-8编码兼容 ──────────────────────────────
_IS_WIN = sys.platform == "win32"
if _IS_WIN:
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

# 终端安全的emoji替代（兼容不支持Unicode的终端）
def _safe(s: str) -> str:
    """将emoji/特殊Unicode替换为ASCII安全字符"""
    replacements = {
        "\U0001f50d": "[*]", "\U0001f30d": "[@]", "\U0001f4e1": "[>]",
        "\U0001f52c": "[#]", "✅": "[OK]", "⚠️": "[!!]",
        "\U0001f6ab": "[X]", "❌": "[X]", "\U0001f4f0": "[i]",
        "\U0001f4cb": "[i]", "\U0001f4c1": "[=]", "\U0001f61e": ":(",
        "\U0001f63f": ":(", "✨": "~~~", "\U0001f3af": "[!]",
        "\U0001f310": "URL:", "\U0001f4e7": "MAIL:", "\U0001f4de": "TEL:",
        "\U0001f4dd": "DESC:",
    }
    for k, v in replacements.items():
        s = s.replace(k, v)
    return s

# ── 日志配置 ──────────────────────────────────────────────
_log_handler = logging.StreamHandler(sys.stdout)
_log_handler.setFormatter(logging.Formatter(
    "%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S"
))
logging.basicConfig(level=logging.INFO, handlers=[_log_handler])
log = logging.getLogger("finder")

# Windows下包装日志以过滤emoji
if _IS_WIN:
    _orig_info = log.info
    _orig_warning = log.warning
    _orig_debug = log.debug
    _orig_error = log.error

    def _safe_info(msg, *args, **kwargs):
        return _orig_info(_safe(str(msg)), *args, **kwargs)
    def _safe_warning(msg, *args, **kwargs):
        return _orig_warning(_safe(str(msg)), *args, **kwargs)
    def _safe_debug(msg, *args, **kwargs):
        return _orig_debug(_safe(str(msg)), *args, **kwargs)
    def _safe_error(msg, *args, **kwargs):
        return _orig_error(_safe(str(msg)), *args, **kwargs)

    log.info = _safe_info
    log.warning = _safe_warning
    log.debug = _safe_debug
    log.error = _safe_error

# ── 请求会话配置 ──────────────────────────────────────────
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:133.0) Gecko/20100101 Firefox/133.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36 Edg/131.0.0.0",
]

# 请求超时（秒）
REQUEST_TIMEOUT = 10
# 请求间隔（秒）
REQUEST_DELAY = (0.5, 1.0)
# 每个搜索引擎最大结果数
MAX_RESULTS_PER_ENGINE = 50
# 并发访问网站时的最大线程数
MAX_SITE_WORKERS = 5
# 默认最大最终结果数
DEFAULT_MAX_RESULTS = 20

# ── 多国家 / 多语言配置 ───────────────────────────────────
# 国家代码 → { google域名, bing市场, ddg地区, 语言代码, 语言名称 }
COUNTRIES = {
    "com": {"google": "google.com",      "bing": "us", "ddg": "us-en",   "lang": "en", "name": "美国"},
    "uk":  {"google": "google.co.uk",    "bing": "gb", "ddg": "uk-en",   "lang": "en", "name": "英国"},
    "de":  {"google": "google.de",       "bing": "de", "ddg": "de-de",   "lang": "de", "name": "德国"},
    "fr":  {"google": "google.fr",       "bing": "fr", "ddg": "fr-fr",   "lang": "fr", "name": "法国"},
    "es":  {"google": "google.es",       "bing": "es", "ddg": "es-es",   "lang": "es", "name": "西班牙"},
    "it":  {"google": "google.it",       "bing": "it", "ddg": "it-it",   "lang": "it", "name": "意大利"},
    "br":  {"google": "google.com.br",   "bing": "br", "ddg": "br-pt",   "lang": "pt", "name": "巴西"},
    "pt":  {"google": "google.pt",       "bing": "pt", "ddg": "pt-pt",   "lang": "pt", "name": "葡萄牙"},
    "nl":  {"google": "google.nl",       "bing": "nl", "ddg": "nl-nl",   "lang": "nl", "name": "荷兰"},
    "jp":  {"google": "google.co.jp",    "bing": "jp", "ddg": "jp-jp",   "lang": "ja", "name": "日本"},
    "kr":  {"google": "google.co.kr",    "bing": "kr", "ddg": "kr-kr",   "lang": "ko", "name": "韩国"},
    "ru":  {"google": "google.ru",       "bing": "ru", "ddg": "ru-ru",   "lang": "ru", "name": "俄罗斯"},
    "mx":  {"google": "google.com.mx",   "bing": "mx", "ddg": "mx-es",   "lang": "es", "name": "墨西哥"},
    "ca":  {"google": "google.ca",       "bing": "ca", "ddg": "ca-en",   "lang": "en", "name": "加拿大"},
    "au":  {"google": "google.com.au",   "bing": "au", "ddg": "au-en",   "lang": "en", "name": "澳大利亚"},
    "in":  {"google": "google.co.in",    "bing": "in", "ddg": "in-en",   "lang": "en", "name": "印度"},
    "tr":  {"google": "google.com.tr",   "bing": "tr", "ddg": "tr-tr",   "lang": "tr", "name": "土耳其"},
    "pl":  {"google": "google.pl",       "bing": "pl", "ddg": "pl-pl",   "lang": "pl", "name": "波兰"},
    "se":  {"google": "google.se",       "bing": "se", "ddg": "se-sv",   "lang": "sv", "name": "瑞典"},
    "ar":  {"google": "google.com.ar",   "bing": "ar", "ddg": "ar-es",   "lang": "es", "name": "阿根廷"},
    "cl":  {"google": "google.cl",       "bing": "cl", "ddg": "cl-es",   "lang": "es", "name": "智利"},
}

# ── 平台网站黑名单 ────────────────────────────────────────
PLATFORM_DOMAINS = {
    # 综合电商
    "amazon", "ebay", "alibaba", "aliexpress", "walmart", "etsy",
    "shopify", "mercadolibre", "rakuten", "flipkart", "jd.com",
    "taobao.com", "tmall.com", "pinduoduo.com",
    # B2B 平台
    "made-in-china.com", "globalsources.com", "tradekey.com",
    "ec21.com", "ecplaza.net", "exportersindia.com", "indiamart.com",
    "thomasnet.com", "kompass.com", "europages.com",
    # 分类信息
    "craigslist.org", "kijiji.ca", "gumtree.com", "olx.",
    # 社交媒体
    "facebook.com", "instagram.com", "twitter.com", "linkedin.com",
    "pinterest.com", "youtube.com", "tiktok.com", "reddit.com",
    # 其他平台
    "wikipedia.org", "quora.com", "answers.com",
}

# 平台网站URL关键词（二级匹配）
PLATFORM_URL_PATTERNS = [
    r"/shop/", r"/store/", r"/product-listing/", r"/marketplace/",
    r"/catalog/products", r"/wholesale/", r"/supplier/",
]

# 新闻/博客域名/URL模式
NEWS_BLOG_PATTERNS = [
    r"\bnews\.", r"\bblog\.", r"\bblogs\.", r"/news/", r"/blog/", r"/blogs/",
    r"medium\.com", r"wordpress\.com", r"blogger\.com", r"tumblr\.com",
    r"forbes\.com", r"bbc\.", r"cnn\.", r"reuters\.", r"bloomberg\.",
    r"techcrunch\.", r"theverge\.", r"wired\.", r"nytimes\.", r"wsj\.com",
    r"prnewswire\.", r"businesswire\.", r"globenewswire\.",
]

# ── 邮箱 / 电话正则 ───────────────────────────────────────
# 通用邮箱正则（支持多语言域名后缀）
EMAIL_RE = re.compile(
    r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}",
    re.IGNORECASE,
)

# 国际电话号码正则（较宽松，支持各种格式）
PHONE_RE = re.compile(
    r"(?:"  # 前缀
    r"(?:\+\d{1,4}[\s.\-]?)?"            # +国家代码
    r"(?:\(?\d{1,4}\)?[\s.\-]?)?"         # 区号
    r")?"
    r"\d{2,4}"                             # 局号
    r"[\s.\-]?"                            # 分隔符
    r"\d{3,4}"                             # 号码中间部分
    r"[\s.\-]?"                            # 分隔符
    r"\d{3,4}"                             # 号码最后部分
)

# 更精确的电话匹配（至少8位数字，有明确格式）
PHONE_STRICT_RE = re.compile(
    r"(?:\+?\d{1,4}[\s.\-]?)?"             # +国家代码
    r"(?:\(?\d{1,4}\)?[\s.\-]?)?"          # 区号
    r"\d{3,4}[\s.\-]?\d{3,4}[\s.\-]?\d{3,4}"  # 号码主体
)

# ── 企业网站特征关键词（多语言）───────────────────────────
COMPANY_INDICATORS = [
    # 英语
    "about us", "contact us", "our products", "our company", "ltd", "inc",
    "llc", "corporation", "manufacturer", "supplier", "factory", "wholesale",
    # 德语
    "über uns", "kontakt", "unsere produkte", "unternehmen", "gmbh", "ag",
    "hersteller", "lieferant", "produkte",
    # 法语
    "à propos", "contactez-nous", "nos produits", "notre entreprise",
    "sarl", "sas", "fabricant", "fournisseur", "produits",
    # 西班牙语
    "sobre nosotros", "contacto", "nuestros productos", "nuestra empresa",
    "s.a.", "s.l.", "fabricante", "proveedor", "productos",
    # 葡萄牙语
    "sobre nós", "contato", "nossos produtos", "nossa empresa",
    "ltda", "fabricante", "fornecedor", "produtos",
    # 意大利语
    "chi siamo", "contatti", "i nostri prodotti", "azienda", "s.p.a.", "s.r.l.",
    "produttore", "fornitore", "prodotti",
    # 荷兰语
    "over ons", "contact", "onze producten", "bedrijf", "b.v.", "n.v.",
    "fabrikant", "leverancier", "producten",
]

# ── 产品栏目/页面关键词（多语言）───────────────────────────
PRODUCT_PAGE_INDICATORS = [
    "product", "products", "produkt", "produkte", "produit", "produits",
    "producto", "productos", "produto", "produtos", "prodotto", "prodotti",
    "producten", "catalog", "catalogue", "katalog", "catálogo", "catálogo",
    "collection", "category", "categories",
]


# ═══════════════════════════════════════════════════════════════
# 工具函数
# ═══════════════════════════════════════════════════════════════

def create_session():
    """创建带随机UA和超时的requests会话"""
    session = requests.Session()
    session.headers.update({
        "User-Agent": random.choice(USER_AGENTS),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "DNT": "1",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
    })
    return session


def polite_delay():
    """礼貌延迟，避免被封"""
    time.sleep(random.uniform(*REQUEST_DELAY))


def extract_domain(url: str) -> str:
    """从URL提取域名"""
    parsed = urllib.parse.urlparse(url)
    domain = parsed.netloc.lower()
    if domain.startswith("www."):
        domain = domain[4:]
    return domain


def normalize_url(url: str) -> str:
    """规范化URL"""
    url = url.strip()
    if not url.startswith(("http://", "https://")):
        url = "https://" + url
    # 去掉尾部多余字符
    url = re.sub(r'[<>"\']+$', '', url)
    return url


def extract_emails(text: str) -> list[str]:
    """从文本中提取邮箱地址，过滤掉常见的假邮箱"""
    emails = EMAIL_RE.findall(text)
    seen = set()
    result = []
    # 过滤掉图片/资源类邮箱、示例邮箱
    skip_patterns = [
        r'\.(png|jpg|jpeg|gif|svg|css|js|woff|ttf|ico)$',
        r'^example@', r'^test@', r'^user@', r'^info@example',
        r'@example\.', r'@test\.', r'@domain\.', r'@email\.',
        r'@yourcompany\.', r'@yourdomain\.', r'@company\.com$',
        r'^noreply@', r'^no-reply@', r'^donotreply@',
    ]
    for email in emails:
        email_lower = email.lower()
        if any(re.search(p, email_lower) for p in skip_patterns):
            continue
        if email_lower not in seen:
            seen.add(email_lower)
            result.append(email)
    return result


def extract_phones(text: str) -> list[str]:
    """从文本中提取电话号码"""
    phones = PHONE_STRICT_RE.findall(text)
    seen = set()
    result = []
    # 过滤明显的假号码
    skip_patterns = [
        r'^0{5,}', r'^1{5,}', r'^2{5,}', r'^3{5,}', r'^4{5,}', r'^5{5,}',
        r'^6{5,}', r'^7{5,}', r'^8{5,}', r'^9{5,}', r'^12345',
        r'\d{4}-\d{2}-\d{2}$',  # 可能是日期
    ]
    for phone in phones:
        phone_clean = phone.strip()
        if not phone_clean:
            continue
        if any(re.search(p, phone_clean) for p in skip_patterns):
            continue
        # 至少8位数字
        digits = re.sub(r'\D', '', phone_clean)
        if len(digits) < 8 or len(digits) > 18:
            continue
        if phone_clean not in seen:
            seen.add(phone_clean)
            result.append(phone_clean)
    return result


def is_platform_site(url: str) -> bool:
    """判断是否为平台型网站（亚马逊、阿里巴巴等）"""
    domain = extract_domain(url)
    # 精确黑名单匹配
    for pd in PLATFORM_DOMAINS:
        if pd in domain:
            return True
    # URL模式匹配
    for pattern in PLATFORM_URL_PATTERNS:
        if re.search(pattern, url, re.IGNORECASE):
            return True
    return False


def is_news_or_blog(url: str) -> bool:
    """判断是否为新闻/博客类网站"""
    domain = extract_domain(url)
    for pattern in NEWS_BLOG_PATTERNS:
        if re.search(pattern, domain, re.IGNORECASE) or re.search(pattern, url, re.IGNORECASE):
            return True
    return False


def has_product_section(html: str, keyword: str) -> bool:
    """判断页面是否包含产品相关内容（非纯新闻/博客）。
    支持多词关键词：只要关键词中的核心词出现足够次数即可。"""
    text_lower = html.lower()[:60000]
    kw_lower = keyword.lower()

    # 拆分关键词（处理多词关键词）
    kw_words = [w.strip() for w in kw_lower.split() if len(w.strip()) > 1]
    # 过滤掉太通用的词
    generic_words = {"the", "a", "an", "and", "or", "for", "with", "of", "in", "on",
                     "to", "is", "are", "was", "were", "be", "been", "it", "at", "by",
                     "de", "la", "le", "les", "des", "et", "en", "el", "los", "las",
                     "da", "do", "das", "dos", "di", "il", "per", "con", "por", "para"}
    core_words = [w for w in kw_words if w not in generic_words]

    # 核心词至少出现1次
    any_match = False
    for word in core_words:
        if word in text_lower:
            any_match = True
            break

    # 完整关键词匹配
    exact_match = kw_lower in text_lower

    if not any_match and not exact_match:
        return False

    # 计算核心词出现次数（至少2个核心词出现）
    core_word_hits = sum(1 for w in core_words if w in text_lower)
    total_occurrences = sum(text_lower.count(w) for w in core_words)

    # 检查是否包含产品相关栏目词
    for indicator in PRODUCT_PAGE_INDICATORS:
        if indicator in text_lower:
            return True

    # 检查是否包含企业特征词
    company_count = sum(1 for ind in COMPANY_INDICATORS if ind in text_lower)
    if company_count >= 2:
        return True

    # 核心词总出现次数 >= 5 或 完整关键词出现 >= 2次
    if total_occurrences >= 5:
        return True

    if text_lower.count(kw_lower) >= 2:
        return True

    # 至少一半核心词出现
    if core_words and core_word_hits >= max(1, len(core_words) // 2):
        return True

    return False


def is_business_site(html: str) -> bool:
    """综合判断是否为企业网站"""
    text_lower = html.lower()[:50000]

    # 检查企业特征
    score = 0
    for indicator in COMPANY_INDICATORS:
        if indicator in text_lower:
            score += 1

    # 检查常见的B2B/B2C链接模式
    business_links = [
        "contact", "/contact", "about", "/about",
        "products", "/products", "produkte", "/produkte",
        "produits", "/produits", "productos", "/productos",
        "produtos", "/produtos",
    ]
    for link in business_links:
        if f'href="{link}"' in text_lower or f"href='{link}'" in text_lower:
            score += 1

    # 检查footer中的版权信息
    if re.search(r'©\s*\d{4}|copyright\s*\d{4}', text_lower, re.IGNORECASE):
        score += 2

    return score >= 3


# ═══════════════════════════════════════════════════════════════
# 搜索引擎模块
# ═══════════════════════════════════════════════════════════════

def search_google(query: str, country_code: str = "com", max_results: int = 30) -> list[dict]:
    """通过谷歌搜索（googlesearch-python + HTTP直搜双重方案）"""
    try:
        from googlesearch import search as gsearch
    except ImportError:
        log.warning("  [!!] googlesearch-python 未安装，跳过Google")
        log.warning("     安装: pip install googlesearch-python")
        return []

    country = COUNTRIES.get(country_code, COUNTRIES["com"])
    lang = country["lang"]
    results = []

    log.info(f"  [*] Google ({country['name']}): lang={lang}")

    try:
        for url in gsearch(
            query,
            num_results=min(max_results, 30),
            lang=lang,
            region=country_code,
            unique=True,
            sleep_interval=3,
            timeout=10,
            advanced=False,
        ):
            if url and url.startswith("http"):
                results.append({
                    "title": "",
                    "url": normalize_url(url),
                    "snippet": "",
                    "engine": "google",
                    "country": country_code,
                })
    except Exception as e:
        pass  # 静默失败，下面尝试直搜

    # 如果googlesearch库没结果，尝试HTTPS直搜
    if not results:
        log.info(f"     googlesearch库无结果，尝试直接HTTPS搜索...")
        try:
            session = create_session()
            params = {
                "q": query, "hl": lang, "num": min(max_results, 30),
                "lr": f"lang_{lang}" if lang != "en" else "",
            }
            url = f"https://www.google.com/search?{urllib.parse.urlencode(params)}"
            resp = session.get(url, timeout=10)
            if resp.status_code == 200 and "captcha" not in resp.text[:1000].lower():
                soup = BeautifulSoup(resp.text, "html.parser")
                for a in soup.select("a[href]"):
                    href = a.get("href", "")
                    if href.startswith("/url?q="):
                        real = urllib.parse.parse_qs(
                            urllib.parse.urlparse("https://google.com" + href).query
                        ).get("q", [""])[0]
                        if real.startswith("http") and "google.com" not in real:
                            results.append({
                                "title": a.get_text(strip=True)[:200],
                                "url": normalize_url(real),
                                "snippet": "",
                                "engine": "google",
                                "country": country_code,
                            })
                        if len(results) >= max_results:
                            break
            else:
                log.warning(f"     Google直接访问被拦截(状态码:{resp.status_code})")
        except Exception:
            pass

    if not results:
        log.warning(f"     [X] Google搜索被限流，建议改用DuckDuckGo引擎")
    else:
        log.info(f"     Google返回 {len(results)} 条")

    polite_delay()
    return results


def search_bing(query: str, country_code: str = "com", max_results: int = 30) -> list[dict]:
    """通过必应搜索（直接爬取，可能被CAPTCHA拦截）"""
    country = COUNTRIES.get(country_code, COUNTRIES["com"])
    cc = country["bing"]
    lang = country["lang"]

    results = []
    session = create_session()
    session.headers["Accept-Language"] = f"{lang}-{cc.upper()},{lang};q=0.9"

    params = {
        "q": query, "cc": cc, "count": min(max_results, 20),
        "setlang": lang, "mkt": f"{lang}-{cc.upper()}",
    }
    url = f"https://www.bing.com/search?{urllib.parse.urlencode(params)}"

    log.info(f"  [*] Bing ({country['name']}): cc={cc}")

    try:
        resp = session.get(url, timeout=12)
        if resp.status_code != 200:
            log.warning(f"     [X] Bing返回状态码 {resp.status_code}")
            return []

        html = resp.text
        if "captcha" in html[:2000].lower() or "Human challenge" in html[:2000]:
            log.warning(f"     [X] Bing要求人机验证，建议改用DuckDuckGo")
            return []

        soup = BeautifulSoup(html, "html.parser")

        # 尝试多种选择器
        for item in soup.select("li.b_algo, .b_algo, ol#b_results > li"):
            link = item.select_one("h2 a[href]") or item.select_one("a[href]")
            if not link:
                continue
            href = link.get("href", "")
            if not href.startswith("http"):
                continue

            title = link.get_text(strip=True)
            snippet = ""
            for s in item.select("p, .b_caption p, .b_algoSlug"):
                snippet = s.get_text(strip=True)
                if snippet:
                    break

            results.append({
                "title": title, "url": normalize_url(href),
                "snippet": snippet, "engine": "bing",
                "country": country_code,
            })
            if len(results) >= max_results:
                break

    except requests.RequestException as e:
        log.warning(f"  [!!] Bing搜索失败: {e}")

    if not results:
        log.warning(f"     [X] Bing未返回结果，建议改用DuckDuckGo引擎")

    polite_delay()
    return results


def search_duckduckgo(query: str, country_code: str = "com", max_results: int = 30) -> list[dict]:
    """通过DuckDuckGo搜索（使用ddgs库，底层调用Bing）。带超时保护。"""
    try:
        from ddgs import DDGS
    except ImportError:
        try:
            from duckduckgo_search import DDGS
        except ImportError:
            log.warning("  [!!] ddgs 库未安装，跳过DuckDuckGo搜索")
            log.warning("     安装方法: pip install ddgs")
            return []

    country = COUNTRIES.get(country_code, COUNTRIES["com"])
    region = country["ddg"]
    results = []
    # DDG 每条结果约1秒，限制数量避免搜索太久
    limit = min(max_results, 25)

    log.info(f"  [*] DDG搜索 ({country['name']}): region={region} 最多{limit}条")

    def _do_search():
        nonlocal results
        try:
            with DDGS() as ddgs:
                for r in ddgs.text(query, region=region, max_results=limit):
                    results.append({
                        "title": r.get("title", ""),
                        "url": normalize_url(r.get("href", "")),
                        "snippet": r.get("body", ""),
                        "engine": "duckduckgo",
                        "country": country_code,
                    })
        except Exception as e:
            log.debug(f"     DDG内部异常: {e}")

    # 用线程+超时防止DDG卡死（某些地区响应慢）
    import threading as _th
    t = _th.Thread(target=_do_search, daemon=True)
    t.start()
    t.join(timeout=30)  # 最多等30秒
    if t.is_alive():
        log.warning(f"     [X] DDG搜索超时(30s)，已放弃等待，请换其他地区或引擎")
        # 线程已daemon，无法强杀，但至少不阻塞

    if results:
        log.info(f"     DDG返回 {len(results)} 条")
    else:
        log.warning(f"     [X] DDG无结果，请尝试: 1)换地区 2)缩短关键词 3)换Bing/Google引擎")

    polite_delay()
    return results


# ── API 搜索（稳定可靠，有免费额度）────────────────────
def search_google_api(query, country_code="com", max_results=30, api_key="", cx=""):
    """Google Custom Search API (免费100次/天)
    申请: https://programmablesearchengine.google.com/ → 创建引擎 → 选'搜索整个网络'"""
    if not api_key or not cx:
        return []
    country = COUNTRIES.get(country_code, COUNTRIES["com"])
    results = []
    log.info(f"  [*] Google API ({country['name']})")
    try:
        params = {"key": api_key, "cx": cx, "q": query, "num": min(max_results, 10)}
        if country["lang"] != "en":
            params["lr"] = f"lang_{country['lang']}"
        resp = requests.get("https://www.googleapis.com/customsearch/v1", params=params, timeout=10)
        if resp.status_code != 200:
            log.warning(f"     Google API {resp.status_code}: {resp.text[:150]}")
            return []
        for item in resp.json().get("items", []):
            results.append({"title": item.get("title", ""), "url": normalize_url(item.get("link", "")),
                            "snippet": item.get("snippet", ""), "engine": "google", "country": country_code})
    except Exception as e:
        log.warning(f"     Google API失败: {e}")
    log.info(f"     Google API返回 {len(results)} 条")
    return results


def search_bing_api(query, country_code="com", max_results=30, api_key=""):
    """Bing Web Search API (免费1000次/月)
    申请: https://portal.azure.com → 创建资源 → 搜索'Bing Search' → 选F0免费层"""
    if not api_key:
        return []
    country = COUNTRIES.get(country_code, COUNTRIES["com"])
    results = []
    mkt = f"{country['lang']}-{country_code.upper()}"
    log.info(f"  [*] Bing API ({country['name']}): mkt={mkt}")
    try:
        headers = {"Ocp-Apim-Subscription-Key": api_key}
        params = {"q": query, "mkt": mkt, "count": min(max_results, 50), "setLang": country["lang"]}
        resp = requests.get("https://api.bing.microsoft.com/v7.0/search", headers=headers, params=params, timeout=10)
        if resp.status_code != 200:
            log.warning(f"     Bing API {resp.status_code}: {resp.text[:150]}")
            return []
        for item in resp.json().get("webPages", {}).get("value", []):
            results.append({"title": item.get("name", ""), "url": normalize_url(item.get("url", "")),
                            "snippet": item.get("snippet", ""), "engine": "bing", "country": country_code})
    except Exception as e:
        log.warning(f"     Bing API失败: {e}")
    log.info(f"     Bing API返回 {len(results)} 条")
    return results


def search_all(
    query: str,
    countries: list[str] = None,
    engines: list[str] = None,
    max_results: int = 30,
    engine_status_callback=None,
    google_api_key: str = "",
    google_cx: str = "",
    bing_api_key: str = "",
) -> list[dict]:
    """统一搜索入口。可选 engine_status_callback(engine_name, country, ok) 报告引擎状态。"""
    if countries is None:
        countries = ["com"]
    if engines is None:
        engines = ["duckduckgo", "google", "bing"]

    all_results = []
    seen_urls = set()
    per_engine = min(max(10, max_results // max(len(engines), 1)), 30)

    for country_code in countries:
        country_name = COUNTRIES.get(country_code, {}).get("name", country_code)
        log.info(f"[@] 搜索地区: {country_name} ({country_code})")

        for engine in engines:
            if engine == "google":
                # 优先API → failback直搜
                if google_api_key and google_cx:
                    found = search_google_api(query, country_code, per_engine, google_api_key, google_cx)
                else:
                    found = []
                if not found:
                    found = search_google(query, country_code, per_engine)
            elif engine == "bing":
                if bing_api_key:
                    found = search_bing_api(query, country_code, per_engine, bing_api_key)
                else:
                    found = []
                if not found:
                    found = search_bing(query, country_code, per_engine)
            elif engine == "duckduckgo":
                found = search_duckduckgo(query, country_code, per_engine)
            else:
                log.warning(f"未知搜索引擎: {engine}")
                continue

            # 报告状态
            if engine_status_callback:
                engine_status_callback(engine, country_code, len(found) > 0)

            # 去重
            for item in found:
                domain = extract_domain(item["url"])
                if domain not in seen_urls:
                    seen_urls.add(domain)
                    all_results.append(item)

            log.info(f"    获得 {len(found)} 条 (累计 {len(all_results)} 条)")

    return all_results


# ═══════════════════════════════════════════════════════════════
# 网站分析模块
# ═══════════════════════════════════════════════════════════════

# ── 联系页面路径（多语言）─────────────────────────────────
CONTACT_PAGE_PATHS = [
    "/contact", "/contact-us", "/contactus", "/contacts",
    "/kontakt", "/kontaktieren",                    # 德语
    "/contactez-nous", "/nous-contacter",           # 法语
    "/contacto", "/contactanos", "/contactar",      # 西班牙语
    "/contato", "/fale-conosco", "/contacte",       # 葡萄牙语
    "/contatti", "/contattaci",                     # 意大利语
    "/contactgegevens",                             # 荷兰语
    "/about", "/about-us", "/uber-uns", "/chi-siamo",
    "/impressum",                                   # 德语法律声明（常有联系方式）
]


def _fetch_page(url: str, session: requests.Session) -> tuple[str | None, str | None]:
    """获取网页HTML内容，返回 (html, final_url)"""
    try:
        resp = session.get(url, timeout=REQUEST_TIMEOUT, allow_redirects=True)
        resp.raise_for_status()

        # 检测编码
        if resp.encoding is None or resp.encoding.lower() == "iso-8859-1":
            match = re.search(rb'charset=["\']?([a-zA-Z0-9\-_]+)', resp.content)
            if match:
                try:
                    resp.encoding = match.group(1).decode("ascii")
                except Exception:
                    resp.encoding = resp.apparent_encoding
            else:
                resp.encoding = resp.apparent_encoding

        return resp.text, resp.url
    except requests.RequestException:
        return None, None


def _extract_info_from_page(html: str) -> tuple[list[str], list[str], str, str]:
    """从单页HTML中提取邮箱、电话、标题、描述"""
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "noscript", "iframe", "svg"]):
        tag.decompose()

    text = soup.get_text(separator=" ", strip=True)

    # 邮箱
    emails = extract_emails(text)
    for mailto in soup.select("a[href^='mailto:']"):
        email = mailto.get("href", "").replace("mailto:", "").split("?")[0].strip()
        if email and "@" in email:
            emails.append(email)
    emails = list(set(emails))

    # 电话
    phones = extract_phones(text)
    for tel_link in soup.select("a[href^='tel:']"):
        phone = tel_link.get("href", "").replace("tel:", "").strip()
        if phone:
            phones.append(phone)
    phones = list(set(phones))

    # 标题 & 描述
    title = soup.title.get_text(strip=True) if soup.title else ""
    desc = ""
    meta_desc = soup.select_one("meta[name='description']")
    if meta_desc:
        desc = meta_desc.get("content", "")[:300]

    return emails, phones, title, desc


def _find_contact_links(html: str, base_url: str) -> list[str]:
    """从主页HTML中查找可能的联系页面链接"""
    soup = BeautifulSoup(html, "html.parser")
    found = []

    # 方法1: 检查常见路径模式
    for path in CONTACT_PAGE_PATHS:
        for a in soup.select(f'a[href*="{path}"]'):
            href = a.get("href", "")
            if href and not href.startswith("#"):
                full_url = urllib.parse.urljoin(base_url, href)
                if full_url not in found:
                    found.append(full_url)

    # 方法2: 检查包含 contact/kontakt 等关键词的链接文本
    contact_words = [
        "contact", "kontakt", "contacto", "contato", "contatti",
        "contactez", "contacter", "reach-us", "get-in-touch",
    ]
    for a in soup.select("a[href]"):
        text = a.get_text(strip=True).lower()
        href = a.get("href", "")
        if href and not href.startswith("#") and not href.startswith("javascript:"):
            if any(w in text for w in contact_words):
                full_url = urllib.parse.urljoin(base_url, href)
                if full_url not in found:
                    found.append(full_url)

    return found[:5]  # 最多检查5个


def fetch_and_analyze(url: str, keyword: str, session: requests.Session = None) -> dict | None:
    """访问网站并提取联系信息（含联系页面自动发现）"""
    if session is None:
        session = create_session()

    domain = extract_domain(url)

    # ── 前置过滤 ──
    if is_platform_site(url):
        log.debug(f"  [平台] 跳过: {domain}")
        return None

    if is_news_or_blog(url):
        log.debug(f"  [新闻/博客] 跳过: {domain}")
        return None

    # ── 访问主页 ──
    html, final_url = _fetch_page(url, session)
    if not html:
        log.debug(f"  [无法访问] {domain}")
        return None

    # ── 判断是否为企业网站 ──
    if not is_business_site(html):
        log.debug(f"  [非企业] {domain}")
        return None

    if not has_product_section(html, keyword):
        log.debug(f"  [无产品栏目] '{keyword}': {domain}")
        return None

    # ── 从主页提取联系方式 ──
    emails, phones, title, desc = _extract_info_from_page(html)

    # ── 如果主页没找到联系方式，尝试联系页面 ──
    if not emails and not phones:
        contact_urls = _find_contact_links(html, final_url or url)

        for contact_url in contact_urls:
            contact_domain = extract_domain(contact_url)
            if contact_domain != domain:
                continue  # 不跳转到其他域名

            polite_delay()
            ch_html, _ = _fetch_page(contact_url, session)
            if not ch_html:
                continue

            c_emails, c_phones, _, _ = _extract_info_from_page(ch_html)
            emails.extend(c_emails)
            phones.extend(c_phones)
            emails = list(set(emails))
            phones = list(set(phones))

            if emails or phones:
                log.debug(f"    从联系页面找到: {contact_url}")
                break

    # ── 构建结果 ──
    if not emails and not phones:
        log.debug(f"  [无联系方式] {domain}")
        return None

    log.info(f"  [找到] {domain} | 邮箱:{len(emails)} 电话:{len(phones)}")

    return {
        "domain": domain,
        "url": url,
        "title": title,
        "description": desc,
        "emails": emails,
        "phones": phones,
        "has_product_page": True,
        "found_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def process_search_results(
    search_results: list[dict],
    keyword: str,
    max_sites: int = DEFAULT_MAX_RESULTS,
    result_callback=None,
    stop_check=None,
) -> list[dict]:
    """处理搜索结果：访问网站、提取信息、过滤。
    result_callback(result, found, visited, total) 实时推送结果
    stop_check() → bool 返回True则停止循环"""
    total = len(search_results)
    log.info(f"\n[>] 开始分析 {total} 个网站（最多获取 {max_sites} 个）...")

    session = create_session()
    results = []
    visited = 0

    for item in search_results:
        if stop_check and stop_check():
            log.info("  [用户停止] 搜索已中止")
            break
        if len(results) >= max_sites:
            break

        url = item["url"]
        domain = extract_domain(url)

        visited += 1
        progress = f"[{visited}/{min(total, max_sites * 3)}]"
        log.info(f"{progress} 正在访问: {domain}")

        result = fetch_and_analyze(url, keyword, session)

        if result:
            result["search_engine"] = item.get("engine", "unknown")
            result["search_country"] = item.get("country", "unknown")
            result["search_snippet"] = item.get("snippet", "")
            results.append(result)

            # 实时回调（用于GUI实时更新）
            if result_callback:
                try:
                    result_callback(result, len(results), visited, total)
                except Exception:
                    pass

        polite_delay()

        # 每10个站点暂停一下
        if visited % 10 == 0:
            time.sleep(1)

    return results


# ═══════════════════════════════════════════════════════════════
# 输出模块
# ═══════════════════════════════════════════════════════════════

def print_results(results: list[dict]):
    """在控制台打印结果"""
    if not results:
        log.info("\n:( 未找到匹配的企业网站。")
        return

    print(_safe("\n" + "=" * 80))
    print(_safe(f"  [!] 共找到 {len(results)} 个潜在客户网站"))
    print("=" * 80)

    for i, r in enumerate(results, 1):
        print(_safe(f"\n{'─' * 60}"))
        print(f"  [{i}] {r['title'][:80]}")
        print(f"  URL: {r['url']}")
        if r.get("emails"):
            print(f"  MAIL: {', '.join(r['emails'][:5])}")
        if r.get("phones"):
            print(f"  TEL: {', '.join(r['phones'][:5])}")
        print(f"  来源: {r.get('search_engine', '?')} | 地区: {r.get('search_country', '?')}")
        if r.get("description"):
            print(f"  DESC: {r['description'][:120]}")


def export_csv(results: list[dict], keyword: str, output_dir: str = "."):
    """导出结果为CSV文件"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_keyword = re.sub(r'[^\w]', '_', keyword)[:30]
    filename = Path(output_dir) / f"leads_{safe_keyword}_{timestamp}.csv"

    fieldnames = [
        "序号", "网站域名", "网址", "页面标题", "邮箱", "电话",
        "搜索引擎", "搜索地区", "搜索摘要", "页面描述", "发现时间",
    ]

    with open(filename, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()

        for i, r in enumerate(results, 1):
            writer.writerow({
                "序号": i,
                "网站域名": r.get("domain", ""),
                "网址": r.get("url", ""),
                "页面标题": r.get("title", ""),
                "邮箱": "; ".join(r.get("emails", [])),
                "电话": "; ".join(r.get("phones", [])),
                "搜索引擎": r.get("search_engine", ""),
                "搜索地区": r.get("search_country", ""),
                "搜索摘要": r.get("search_snippet", ""),
                "页面描述": r.get("description", ""),
                "发现时间": r.get("found_at", ""),
            })

    log.info(f"\n📁 结果已保存到: {filename.resolve()}")
    return filename


def export_excel(results: list[dict], keyword: str, output_dir: str = "."):
    """导出结果为Excel文件（需要openpyxl）"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        log.warning("openpyxl 未安装，跳过Excel导出。安装: pip install openpyxl")
        return None

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_keyword = re.sub(r'[^\w]', '_', keyword)[:30]
    filename = Path(output_dir) / f"leads_{safe_keyword}_{timestamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "客户线索"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = ["序号", "网站域名", "网址", "页面标题", "邮箱", "电话",
               "搜索引擎", "搜索地区", "搜索摘要", "发现时间"]
    col_widths = [6, 25, 45, 40, 35, 30, 12, 10, 50, 18]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = width

    # 数据行
    for i, r in enumerate(results, 1):
        row_data = [
            i,
            r.get("domain", ""),
            r.get("url", ""),
            r.get("title", ""),
            "; ".join(r.get("emails", [])),
            "; ".join(r.get("phones", [])),
            r.get("search_engine", ""),
            r.get("search_country", ""),
            r.get("search_snippet", ""),
            r.get("found_at", ""),
        ]
        for col, value in enumerate(row_data, 1):
            ws.cell(row=i + 1, column=col, value=value)

    # 冻结首行
    ws.freeze_panes = "A2"
    # 自动筛选
    ws.auto_filter.ref = ws.dimensions

    wb.save(filename)
    log.info(f"📁 Excel结果已保存到: {filename.resolve()}")
    return filename


# ═══════════════════════════════════════════════════════════════
# 主入口
# ═══════════════════════════════════════════════════════════════

def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="🔍 外贸客户号码查找工具 - 通过搜索引擎发现企业客户联系信息",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python number_finder.py "LED lights"                      # 默认谷歌.com搜索
  python number_finder.py "LED lights" -c de                # 谷歌德国站
  python number_finder.py "acero inoxidable" -c es -e bing  # 必应西班牙站
  python number_finder.py "moteur électrique" -c fr de uk   # 多国搜索
  python number_finder.py "stahl rohr" -c de -n 50 -o excel # 导出Excel
  python number_finder.py "motor elétrico" -c br pt -e google bing  # 多引擎
        """,
    )

    parser.add_argument("keyword", nargs="?", default=None, help="产品关键词（支持小语种）")
    parser.add_argument(
        "-c", "--countries", nargs="+", default=["com"],
        help="搜索国家代码（默认: com）。可选: com, uk, de, fr, es, it, br, pt, nl, jp, kr, ru, mx, ca, au, in, tr, pl, se, ar, cl",
    )
    parser.add_argument(
        "-e", "--engines", nargs="+", default=["duckduckgo", "google", "bing"],
        help="搜索引擎（默认: duckduckgo google bing）。推荐优先使用 duckduckgo（无需API、较稳定）",
    )
    parser.add_argument(
        "-n", "--max-results", type=int, default=DEFAULT_MAX_RESULTS,
        help=f"最大输出结果数（默认: {DEFAULT_MAX_RESULTS}）",
    )
    parser.add_argument(
        "-o", "--output", choices=["csv", "excel", "both"], default="csv",
        help="输出格式（默认: csv）",
    )
    parser.add_argument(
        "--output-dir", default=".",
        help="输出目录（默认: 当前目录）",
    )
    parser.add_argument(
        "-v", "--verbose", action="store_true",
        help="显示详细日志",
    )
    parser.add_argument(
        "--list-countries", action="store_true",
        help="列出支持的国家/地区代码",
    )

    args = parser.parse_args()

    # 列出国家
    if args.list_countries:
        print("\n支持的国家/地区代码:")
        print("-" * 50)
        for code, info in COUNTRIES.items():
            print(f"  {code:6s} → {info['name']:10s}  Google: {info['google']:20s}  语言: {info['lang']}")
        return

    # 关键词必填
    if not args.keyword:
        parser.error("请提供产品关键词。示例: python number_finder.py \"LED lights\"")

    # 验证国家代码
    for cc in args.countries:
        if cc not in COUNTRIES:
            log.error(f"不支持的国家代码: {cc}")
            log.info(f"支持的国家代码: {', '.join(COUNTRIES.keys())}")
            sys.exit(1)

    # 验证搜索引擎
    valid_engines = {"google", "bing", "duckduckgo"}
    for eng in args.engines:
        if eng not in valid_engines:
            log.error(f"不支持的搜索引擎: {eng}")
            log.info(f"支持的搜索引擎: {', '.join(valid_engines)}")
            sys.exit(1)

    # 日志级别
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    # 打印配置
    print("\n" + "=" * 60)
    print(_safe("  [*] 外贸客户号码查找工具"))
    print("=" * 60)
    print(f"  关键词: {args.keyword}")
    print(f"  搜索地区: {', '.join(args.countries)}")
    print(f"  搜索引擎: {', '.join(args.engines)}")
    print(f"  最大结果: {args.max_results}")
    print(f"  输出格式: {args.output}")
    print("=" * 60)

    # ── 第一步：搜索 ──
    log.info("\n📡 第一步：搜索引擎搜索...")
    search_results = search_all(
        query=args.keyword,
        countries=args.countries,
        engines=args.engines,
        max_results=args.max_results * 3,  # 搜索3倍结果，后续过滤
    )
    log.info(f"✅ 搜索完成，共获得 {len(search_results)} 条去重结果")

    if not search_results:
        log.error("没有搜索到任何结果，请检查网络或更换关键词。")
        sys.exit(1)

    # ── 第二步：分析网站 ──
    log.info("\n🔬 第二步：访问网站、提取联系信息...")
    results = process_search_results(
        search_results=search_results,
        keyword=args.keyword,
        max_sites=args.max_results,
    )
    log.info(f"✅ 网站分析完成，找到 {len(results)} 个潜在客户")

    # ── 第三步：输出 ──
    print_results(results)

    if results:
        if args.output in ("csv", "both"):
            export_csv(results, args.keyword, args.output_dir)
        if args.output in ("excel", "both"):
            export_excel(results, args.keyword, args.output_dir)

    print(_safe("\n~~~ 任务完成! ~~~"))


if __name__ == "__main__":
    main()
